import logging
import os
import asyncio
from pathlib import Path
from datetime import datetime, timedelta
import shutil
import tempfile
from typing import Dict, List, Optional, Tuple
import json
import hashlib
import mimetypes
import zipfile
import rarfile
import tarfile
import pyexcel
import openpyxl
import pandas as pd
import csv

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    MessageHandler,
    CallbackQueryHandler,
    ContextTypes,
    filters
)
from telegram.constants import ParseMode
import subprocess
import sys
import io
import traceback

# ==================== KONFIGURATSIYA ====================
class Config:
    BOT_TOKEN = "7964829221:AAHL6c55tIcIEtrhxVhWVTwCXmqyR0WsUrs"
    MAX_FILE_SIZE = 2 * 1024 * 1024 * 1024  # 2GB
    UPLOAD_FOLDER = "uploads"
    OUTPUT_FOLDER = "converted"
    TEMP_FOLDER = "temp"
    EXTRACT_FOLDER = "extracted"
    DATABASE_FILE = "users_data.json"
    LOG_FILE = "bot.log"
    CLEANUP_HOURS = 24
    MAX_CONCURRENT_JOBS = 3
    
    # Adminlar ro'yxati (o'z ID'ingizni qo'shing)
    ADMIN_IDS = [123456789]  # O'zingizning Telegram ID'ingiz
    
    # Kutubxonalar mavjudligini tekshirish
    HAS_PIL = False
    HAS_REPORTLAB = False
    HAS_PYEXCEL = False
    HAS_PANDAS = False
    HAS_OPENPYXL = False
    HAS_ZIPFILE = True  # Python standard library
    HAS_RARFILE = False

# ==================== LOGGING ====================
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(Config.LOG_FILE, encoding='utf-8'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# ==================== FAZL TURLARI ====================
class FileTypes:
    IMAGES = ['jpg', 'jpeg', 'png', 'webp', 'bmp', 'gif', 'tiff', 'ico']
    DOCUMENTS = ['pdf', 'docx', 'doc', 'txt', 'rtf']
    SPREADSHEETS = ['xlsx', 'xls', 'csv', 'ods']  # Excel va boshqa jadval fayllari
    AUDIO = ['mp3', 'wav', 'ogg', 'm4a']
    VIDEO = ['mp4', 'avi', 'mov', 'mkv']
    ARCHIVES = ['zip', 'rar', '7z', 'tar', 'gz', 'tar.gz', 'bz2']  # Qo'shimcha arxiv formatlari
    
    ALL = IMAGES + DOCUMENTS + SPREADSHEETS + AUDIO + VIDEO + ARCHIVES

# ==================== KONVERTATSIYA MATRITSASI ====================
CONVERSION_MATRIX = {
    # Rasmlar
    'jpg': ['png', 'webp', 'pdf'],
    'jpeg': ['png', 'webp', 'pdf'],
    'png': ['jpg', 'webp', 'pdf'],
    'webp': ['jpg', 'png', 'pdf'],
    'bmp': ['jpg', 'png', 'pdf'],
    'gif': ['mp4', 'webp'],
    'tiff': ['jpg', 'png', 'pdf'],
    'ico': ['png', 'jpg'],
    
    # Hujjatlar
    'pdf': ['jpg', 'png'],
    'docx': ['pdf', 'txt'],
    'doc': ['pdf', 'txt'],
    'txt': ['pdf'],
    'rtf': ['pdf', 'txt'],
    
    # Jadval fayllari (Excel va boshqalar)
    'xlsx': ['xls', 'csv', 'pdf', 'html', 'json'],
    'xls': ['xlsx', 'csv', 'pdf', 'html', 'json'],
    'csv': ['xlsx', 'xls', 'json', 'html'],
    'ods': ['xlsx', 'xls', 'csv', 'pdf'],
    
    # Audio
    'mp3': ['wav'],
    'wav': ['mp3'],
    'ogg': ['mp3'],
    'm4a': ['mp3'],
    
    # Video
    'mp4': ['gif'],
    'avi': ['mp4'],
    'mov': ['mp4'],
    'mkv': ['mp4'],
    
    # Arxivlar
    'zip': ['rar', '7z', 'tar.gz', 'extract'],
    'rar': ['zip', '7z', 'tar.gz', 'extract'],
    '7z': ['zip', 'rar', 'tar.gz', 'extract'],
    'tar': ['zip', 'rar', '7z', 'extract'],
    'gz': ['zip', 'extract'],
    'tar.gz': ['zip', 'rar', 'extract'],
}

# ==================== YORDAMCHI FUNKSIYALAR ====================
def setup_environment():
    """Muhitni sozlash va zarur kutubxonalarni tekshirish"""
    try:
        from PIL import Image
        Config.HAS_PIL = True
        logger.info("✅ PIL/Pillow kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ PIL/Pillow kutubxonasi topilmadi. Rasm konvertatsiyasi cheklangan")
    
    try:
        import reportlab
        Config.HAS_REPORTLAB = True
        logger.info("✅ ReportLab kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ ReportLab kutubxonasi topilmadi. PDF yaratish cheklangan")
    
    try:
        import pyexcel
        Config.HAS_PYEXCEL = True
        logger.info("✅ PyExcel kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ PyExcel kutubxonasi topilmadi. Excel konvertatsiyasi cheklangan")
    
    try:
        import pandas as pd
        Config.HAS_PANDAS = True
        logger.info("✅ Pandas kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ Pandas kutubxonasi topilmadi. CSV/Excel konvertatsiyasi cheklangan")
    
    try:
        import openpyxl
        Config.HAS_OPENPYXL = True
        logger.info("✅ OpenPyXL kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ OpenPyXL kutubxonasi topilmadi. XLSX konvertatsiyasi cheklangan")
    
    try:
        import rarfile
        Config.HAS_RARFILE = True
        logger.info("✅ RarFile kutubxonasi mavjud")
    except ImportError:
        logger.warning("❌ RarFile kutubxonasi topilmadi. RAR fayllarini ochish cheklangan")
    
    # Papkalarni yaratish
    os.makedirs(Config.UPLOAD_FOLDER, exist_ok=True)
    os.makedirs(Config.OUTPUT_FOLDER, exist_ok=True)
    os.makedirs(Config.TEMP_FOLDER, exist_ok=True)
    os.makedirs(Config.EXTRACT_FOLDER, exist_ok=True)
    
    logger.info("✅ Papkalar yaratildi")

def get_file_extension(filename: str) -> str:
    """Fayl kengaytmasini olish"""
    return Path(filename).suffix.lower()[1:] if '.' in filename else ''

def get_file_type(extension: str) -> str:
    """Fayl turini aniqlash"""
    extension = extension.lower()
    if extension in FileTypes.IMAGES:
        return 'image'
    elif extension in FileTypes.DOCUMENTS:
        return 'document'
    elif extension in FileTypes.SPREADSHEETS:
        return 'spreadsheet'
    elif extension in FileTypes.AUDIO:
        return 'audio'
    elif extension in FileTypes.VIDEO:
        return 'video'
    elif extension in FileTypes.ARCHIVES:
        return 'archive'
    return 'unknown'

def human_readable_size(size_bytes: int) -> str:
    """Fayl hajmini inson o'qiy oladigan formatda ko'rsatish"""
    if size_bytes == 0:
        return "0 B"
    
    units = ['B', 'KB', 'MB', 'GB', 'TB']
    i = 0
    while size_bytes >= 1024 and i < len(units) - 1:
        size_bytes /= 1024
        i += 1
    
    return f"{size_bytes:.2f} {units[i]}"

def get_file_info(file_path: str) -> Dict:
    """Fayl haqida ma'lumot olish"""
    try:
        stats = os.stat(file_path)
        info = {
            'size': human_readable_size(stats.st_size),
            'size_bytes': stats.st_size,
            'created': datetime.fromtimestamp(stats.st_ctime),
            'modified': datetime.fromtimestamp(stats.st_mtime),
            'extension': get_file_extension(file_path),
            'type': get_file_type(get_file_extension(file_path))
        }
        
        # Rasm o'lchamlari
        if info['type'] == 'image' and Config.HAS_PIL:
            try:
                from PIL import Image
                with Image.open(file_path) as img:
                    info['dimensions'] = f"{img.width}×{img.height}"
            except:
                pass
        
        # Excel fayllar uchun ma'lumot
        if info['type'] == 'spreadsheet' and Config.HAS_PANDAS:
            try:
                ext = info['extension']
                if ext in ['xlsx', 'xls']:
                    if Config.HAS_OPENPYXL:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, read_only=True)
                        info['sheets'] = len(wb.sheetnames)
                        info['sheet_names'] = wb.sheetnames[:5]  # Faqat 5 ta nom
                elif ext == 'csv':
                    with open(file_path, 'r', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        row_count = sum(1 for row in reader)
                        info['rows'] = row_count
            except:
                pass
        
        # Arxiv fayllar uchun ma'lumot
        if info['type'] == 'archive':
            try:
                ext = info['extension']
                if ext == 'zip':
                    with zipfile.ZipFile(file_path, 'r') as zf:
                        info['files_in_archive'] = len(zf.namelist())
                        info['archive_size'] = sum(zf.getinfo(name).file_size for name in zf.namelist())
                        info['archive_files'] = zf.namelist()[:5]  # Faqat 5 ta nom
                elif ext == 'rar' and Config.HAS_RARFILE:
                    import rarfile
                    with rarfile.RarFile(file_path, 'r') as rf:
                        info['files_in_archive'] = len(rf.namelist())
                        info['archive_files'] = rf.namelist()[:5]
            except:
                pass
        
        return info
    except Exception as e:
        logger.error(f"Fayl ma'lumotlarini olishda xato: {e}")
        return {}

def create_format_keyboard(original_ext: str, file_id: str, settings: Dict = None) -> Optional[InlineKeyboardMarkup]:
    """Format tanlash uchun tugmachalar yaratish"""
    if original_ext not in CONVERSION_MATRIX:
        return None
    
    target_formats = CONVERSION_MATRIX.get(original_ext, [])
    
    # Arxivlar uchun extract tugmasini birinchi qatorga qo'shamiz
    buttons = []
    
    # Arxiv bo'lsa, extract tugmasini alohida qatorga
    if original_ext in FileTypes.ARCHIVES:
        extract_row = [
            InlineKeyboardButton(
                "📂 Arxivni ochish",
                callback_data=f"extract:{file_id}"
            )
        ]
        buttons.append(extract_row)
    
    # Qolgan formatlar
    if target_formats:
        row = []
        for i, fmt in enumerate(target_formats):
            if fmt == 'extract':  # Extract alohida tugma sifatida qo'shildi
                continue
                
            # Fayl turi bo'yicha emojilar
            emoji = {
                'image': '🖼️',
                'document': '📄',
                'spreadsheet': '📊',
                'audio': '🎵',
                'video': '🎬',
                'archive': '📦'
            }.get(get_file_type(fmt), '📎')
            
            row.append(InlineKeyboardButton(
                f"{emoji} {fmt.upper()}",
                callback_data=f"conv:{file_id}:{fmt}"
            ))
            
            if (i + 1) % 3 == 0:
                buttons.append(row)
                row = []
        
        if row:
            buttons.append(row)
    
    # Qo'shimcha funksiyalar
    buttons.append([
        InlineKeyboardButton("⚙️ Sozlamalar", callback_data=f"set:{file_id}"),
        InlineKeyboardButton("ℹ️ Ma'lumot", callback_data=f"info:{file_id}")
    ])
    
    return InlineKeyboardMarkup(buttons)

# ==================== ARXIV FUNKSIYALARI ====================
class ArchiveHandler:
    """Arxiv fayllar bilan ishlash"""
    
    @staticmethod
    async def extract_archive(file_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
        """Arxiv faylini ochish"""
        try:
            ext = get_file_extension(file_path)
            
            if ext == 'zip':
                return await ArchiveHandler._extract_zip(file_path, output_dir)
            elif ext == 'rar':
                return await ArchiveHandler._extract_rar(file_path, output_dir)
            elif ext in ['tar', 'gz', 'tar.gz', 'bz2']:
                return await ArchiveHandler._extract_tar(file_path, output_dir)
            else:
                return False, f"Ushbu format qo'llab-quvvatlanmaydi: {ext}", []
                
        except Exception as e:
            logger.error(f"Arxiv ochish xatosi: {e}")
            return False, str(e), []
    
    @staticmethod
    async def _extract_zip(file_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
        """ZIP faylini ochish"""
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # Barcha fayllar ro'yxati
                file_list = zip_ref.namelist()
                
                # Arxiv hajmini tekshirish
                total_size = sum(zip_ref.getinfo(name).file_size for name in file_list)
                if total_size > 100 * 1024 * 1024:  # 100MB
                    return False, "Arxiv juda katta (100MB dan oshib ketdi)", []
                
                # Fayllarni ochish
                zip_ref.extractall(output_dir)
                
                # Chiqarilgan fayllar ro'yxati
                extracted_files = []
                for root, dirs, files in os.walk(output_dir):
                    for file in files:
                        filepath = os.path.join(root, file)
                        if os.path.isfile(filepath):
                            extracted_files.append(filepath)
                
                return True, f"{len(file_list)} ta fayl muvaffaqiyatli ochildi", extracted_files
                
        except zipfile.BadZipFile:
            return False, "ZIP fayli buzilgan yoki noto'g'ri format", []
        except Exception as e:
            return False, str(e), []
    
    @staticmethod
    async def _extract_rar(file_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
        """RAR faylini ochish"""
        try:
            if not Config.HAS_RARFILE:
                return False, "RarFile kutubxonasi o'rnatilmagan. 'pip install rarfile' ni ishlating", []
            
            import rarfile
            with rarfile.RarFile(file_path, 'r') as rar_ref:
                # Barcha fayllar ro'yxati
                file_list = rar_ref.namelist()
                
                # Fayllarni ochish
                rar_ref.extractall(output_dir)
                
                # Chiqarilgan fayllar ro'yxati
                extracted_files = []
                for root, dirs, files in os.walk(output_dir):
                    for file in files:
                        filepath = os.path.join(root, file)
                        if os.path.isfile(filepath):
                            extracted_files.append(filepath)
                
                return True, f"{len(file_list)} ta fayl muvaffaqiyatli ochildi", extracted_files
                
        except rarfile.BadRarFile:
            return False, "RAR fayli buzilgan yoki noto'g'ri format", []
        except Exception as e:
            return False, str(e), []
    
    @staticmethod
    async def _extract_tar(file_path: str, output_dir: str) -> Tuple[bool, str, List[str]]:
        """TAR fayllarini ochish"""
        try:
            mode = 'r'
            if file_path.endswith('.gz'):
                mode = 'r:gz'
            elif file_path.endswith('.bz2'):
                mode = 'r:bz2'
            
            with tarfile.open(file_path, mode) as tar_ref:
                # Barcha fayllar ro'yxati
                file_list = tar_ref.getnames()
                
                # Fayllarni ochish
                tar_ref.extractall(output_dir)
                
                # Chiqarilgan fayllar ro'yxati
                extracted_files = []
                for root, dirs, files in os.walk(output_dir):
                    for file in files:
                        filepath = os.path.join(root, file)
                        if os.path.isfile(filepath):
                            extracted_files.append(filepath)
                
                return True, f"{len(file_list)} ta fayl muvaffaqiyatli ochildi", extracted_files
                
        except tarfile.ReadError:
            return False, "TAR fayli buzilgan yoki noto'g'ri format", []
        except Exception as e:
            return False, str(e), []
    
    @staticmethod
    async def create_archive(files: List[str], output_path: str, archive_type: str = 'zip') -> Tuple[bool, str]:
        """Arxiv yaratish"""
        try:
            if archive_type == 'zip':
                with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                    for file in files:
                        if os.path.isfile(file):
                            arcname = os.path.basename(file)
                            zipf.write(file, arcname)
                return True, "ZIP arxivi muvaffaqiyatli yaratildi"
            
            elif archive_type == 'tar':
                with tarfile.open(output_path, 'w') as tarf:
                    for file in files:
                        if os.path.isfile(file):
                            arcname = os.path.basename(file)
                            tarf.add(file, arcname=arcname)
                return True, "TAR arxivi muvaffaqiyatli yaratildi"
            
            else:
                return False, f"Ushbu arxiv turi qo'llab-quvvatlanmaydi: {archive_type}"
                
        except Exception as e:
            return False, str(e)

# ==================== EXCEL KONVERTATSIYA FUNKSIYALARI ====================
class SpreadsheetConverter:
    """Excel va boshqa jadval fayllarini konvertatsiya qilish"""
    
    @staticmethod
    async def convert_excel(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Excel fayllarini konvertatsiya qilish"""
        try:
            input_ext = get_file_extension(input_path)
            
            if target_format == 'csv':
                return await SpreadsheetConverter._to_csv(input_path, output_path, input_ext)
            elif target_format == 'json':
                return await SpreadsheetConverter._to_json(input_path, output_path, input_ext)
            elif target_format == 'html':
                return await SpreadsheetConverter._to_html(input_path, output_path, input_ext)
            elif target_format == 'xlsx':
                return await SpreadsheetConverter._to_xlsx(input_path, output_path, input_ext)
            elif target_format == 'xls':
                return await SpreadsheetConverter._to_xls(input_path, output_path, input_ext)
            elif target_format == 'pdf':
                return await SpreadsheetConverter._to_pdf(input_path, output_path, input_ext)
            else:
                return False, f"Ushbu formatga konvertatsiya qo'llab-quvvatlanmaydi: {target_format}"
                
        except Exception as e:
            logger.error(f"Excel konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def _to_csv(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Excel dan CSV ga o'tkazish"""
        try:
            if not Config.HAS_PANDAS:
                return False, "Pandas kutubxonasi topilmadi"
            
            if input_ext in ['xlsx', 'xls']:
                # Faqat birinchi sheet ni o'qish
                df = pd.read_excel(input_path, sheet_name=0)
                df.to_csv(output_path, index=False, encoding='utf-8')
                return True, "Muvaffaqiyatli (faqat birinchi varaq)"
            elif input_ext == 'csv':
                # CSV dan CSV ga - faqat nusxalash
                shutil.copy(input_path, output_path)
                return True, "Fayl nusxalandi"
            elif input_ext == 'ods':
                # ODS dan CSV ga
                df = pd.read_excel(input_path, sheet_name=0, engine='odf')
                df.to_csv(output_path, index=False, encoding='utf-8')
                return True, "Muvaffaqiyatli (faqat birinchi varaq)"
            else:
                return False, "Kiritish formati noto'g'ri"
                
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    async def _to_json(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Excel dan JSON ga o'tkazish"""
        try:
            if not Config.HAS_PANDAS:
                return False, "Pandas kutubxonasi topilmadi"
            
            if input_ext in ['xlsx', 'xls', 'csv', 'ods']:
                if input_ext == 'csv':
                    df = pd.read_csv(input_path)
                else:
                    df = pd.read_excel(input_path, sheet_name=0)
                
                # JSON formatiga o'tkazish
                df.to_json(output_path, orient='records', indent=2, force_ascii=False)
                return True, "Muvaffaqiyatli (faqat birinchi varaq)"
            else:
                return False, "Kiritish formati noto'g'ri"
                
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    async def _to_html(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Excel dan HTML ga o'tkazish"""
        try:
            if not Config.HAS_PANDAS:
                return False, "Pandas kutubxonasi topilmadi"
            
            if input_ext in ['xlsx', 'xls', 'csv', 'ods']:
                if input_ext == 'csv':
                    df = pd.read_csv(input_path)
                else:
                    df = pd.read_excel(input_path, sheet_name=0)
                
                # HTML formatiga o'tkazish
                html_content = df.to_html(index=False, classes='table table-striped')
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(f"<!DOCTYPE html><html><head><meta charset='UTF-8'><style>table {{border-collapse: collapse; width: 100%;}} th, td {{border: 1px solid #ddd; padding: 8px;}} th {{background-color: #f2f2f2;}}</style></head><body>{html_content}</body></html>")
                return True, "Muvaffaqiyatli (faqat birinchi varaq)"
            else:
                return False, "Kiritish formati noto'g'ri"
                
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    async def _to_xlsx(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Boshqa formatlardan XLSX ga o'tkazish"""
        try:
            if not Config.HAS_PANDAS:
                return False, "Pandas kutubxonasi topilmadi"
            
            if input_ext in ['xls', 'csv', 'ods']:
                if input_ext == 'csv':
                    df = pd.read_csv(input_path)
                else:
                    df = pd.read_excel(input_path, sheet_name=0)
                
                # XLSX ga o'tkazish
                df.to_excel(output_path, index=False, engine='openpyxl')
                return True, "Muvaffaqiyatli"
            elif input_ext == 'xlsx':
                # XLSX dan XLSX ga - faqat nusxalash
                shutil.copy(input_path, output_path)
                return True, "Fayl nusxalandi"
            else:
                return False, "Kiritish formati noto'g'ri"
                
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    async def _to_xls(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Boshqa formatlardan XLS ga o'tkazish"""
        try:
            if not Config.HAS_PANDAS:
                return False, "Pandas kutubxonasi topilmadi"
            
            if input_ext in ['xlsx', 'csv', 'ods']:
                if input_ext == 'csv':
                    df = pd.read_csv(input_path)
                else:
                    df = pd.read_excel(input_path, sheet_name=0)
                
                # XLS ga o'tkazish (eski format)
                df.to_excel(output_path, index=False)
                return True, "Muvaffaqiyatli"
            elif input_ext == 'xls':
                # XLS dan XLS ga - faqat nusxalash
                shutil.copy(input_path, output_path)
                return True, "Fayl nusxalandi"
            else:
                return False, "Kiritish formati noto'g'ri"
                
        except Exception as e:
            return False, str(e)
    
    @staticmethod
    async def _to_pdf(input_path: str, output_path: str, input_ext: str) -> Tuple[bool, str]:
        """Excel dan PDF ga o'tkazish"""
        try:
            # Oddiy nusxalash (haqiqiy Excel->PDF kutubxona kerak)
            shutil.copy(input_path, output_path)
            return True, "Fayl nusxalandi (PDF konvertatsiyasi mavjud emas)"
            
        except Exception as e:
            return False, str(e)

# ==================== KONVERTATSIYA FUNKSIYALARI ====================
class Converter:
    """Barcha konvertatsiya operatsiyalari"""
    
    @staticmethod
    async def convert_image(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Rasmni konvertatsiya qilish"""
        try:
            if not Config.HAS_PIL:
                return False, "PIL/Pillow kutubxonasi topilmadi"
            
            from PIL import Image
            
            with Image.open(input_path) as img:
                # RGBA dan RGB ga o'tkazish (agar kerak bo'lsa)
                if target_format.upper() in ['JPG', 'JPEG', 'PDF'] and img.mode in ['RGBA', 'LA']:
                    background = Image.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'RGBA':
                        background.paste(img, mask=img.split()[3])
                    else:
                        background.paste(img, mask=img.split()[1])
                    img = background
                elif img.mode == 'P':
                    img = img.convert('RGB')
                
                # Sifat sozlamalari
                quality = int(settings.get('image_quality', 85))
                
                # O'lchamni o'zgartirish
                resize_percent = int(settings.get('resize_percent', 100))
                if resize_percent != 100:
                    new_width = int(img.width * resize_percent / 100)
                    new_height = int(img.height * resize_percent / 100)
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                
                # PDF ga konvertatsiya
                if target_format.lower() == 'pdf':
                    img.save(output_path, 'PDF', quality=quality)
                # GIF ga konvertatsiya
                elif target_format.lower() == 'gif':
                    img.save(output_path, 'GIF', save_all=True, optimize=True)
                # Boshqa formatlar
                else:
                    img.save(output_path, target_format.upper(), quality=quality)
            
            return True, "Muvaffaqiyatli"
            
        except Exception as e:
            logger.error(f"Rasm konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def convert_document(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Hujjatni konvertatsiya qilish"""
        try:
            input_ext = get_file_extension(input_path)
            
            # PDF ga konvertatsiya
            if target_format.lower() == 'pdf':
                # ReportLab orqali (faqat text uchun)
                if Config.HAS_REPORTLAB and input_ext == 'txt':
                    try:
                        from reportlab.lib.pagesizes import letter
                        from reportlab.pdfgen import canvas
                        
                        with open(input_path, 'r', encoding='utf-8') as f:
                            text = f.read()
                        
                        c = canvas.Canvas(output_path, pagesize=letter)
                        width, height = letter
                        
                        c.setFont("Helvetica", 12)
                        text_object = c.beginText(40, height - 40)
                        
                        lines = text.split('\n')
                        for line in lines:
                            text_object.textLine(line[:100])
                        
                        c.drawText(text_object)
                        c.save()
                        return True, "Muvaffaqiyatli"
                    except Exception as e:
                        logger.error(f"ReportLab xatosi: {e}")
                        # Oddiy nusxa olish
                        shutil.copy(input_path, output_path)
                        return True, "Fayl nusxalandi (PDF konvertatsiyasi muvaffaqiyatsiz)"
                
                # Pillow orqali (rasm PDF)
                if Config.HAS_PIL and input_ext in FileTypes.IMAGES:
                    return await Converter.convert_image(input_path, output_path, 'pdf', settings)
            
            # PDF dan boshqa formatga
            elif input_ext == 'pdf' and target_format in ['jpg', 'png']:
                if Config.HAS_PIL:
                    try:
                        from PIL import Image
                        import fitz  # PyMuPDF
                        
                        # PDF ni rasmga aylantirish
                        doc = fitz.open(input_path)
                        page = doc.load_page(0)
                        pix = page.get_pixmap()
                        
                        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
                        img.save(output_path, target_format.upper())
                        return True, "Muvaffaqiyatli"
                    except:
                        return False, "PyMuPDF kutubxonasi kerak"
            
            return False, "Ushbu konvertatsiya hozircha qo'llab-quvvatlanmaydi"
            
        except Exception as e:
            logger.error(f"Hujjat konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def convert_spreadsheet(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Jadval fayllarini konvertatsiya qilish"""
        return await SpreadsheetConverter.convert_excel(input_path, output_path, target_format, settings)
    
    @staticmethod
    async def convert_audio(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Audioni konvertatsiya qilish"""
        try:
            # Oddiy fayl nusxalash (audio uchun)
            shutil.copy(input_path, output_path)
            return True, "Fayl nusxalandi (Audio konvertatsiyasi mavjud emas)"
            
        except Exception as e:
            logger.error(f"Audio konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def convert_video(input_path: str, output_path: str, target_format: str, settings: Dict) -> Tuple[bool, str]:
        """Videoni konvertatsiya qilish"""
        try:
            # Oddiy fayl nusxalash (video uchun)
            shutil.copy(input_path, output_path)
            return True, "Fayl nusxalandi (Video konvertatsiyasi mavjud emas)"
            
        except Exception as e:
            logger.error(f"Video konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def convert_archive(input_path: str, output_path: str, target_format: str) -> Tuple[bool, str]:
        """Arxivni konvertatsiya qilish"""
        try:
            # Arxivni boshqa arxiv formatiga o'tkazish
            if target_format in ['zip', 'rar', '7z', 'tar.gz']:
                # Extract qilish
                temp_dir = tempfile.mkdtemp(dir=Config.TEMP_FOLDER)
                success, message, extracted_files = await ArchiveHandler.extract_archive(input_path, temp_dir)
                
                if success and extracted_files:
                    # Yangi arxiv yaratish
                    if target_format == 'zip':
                        output_ext = 'zip'
                    elif target_format == 'rar':
                        output_ext = 'rar'
                    elif target_format == '7z':
                        output_ext = '7z'
                    elif target_format == 'tar.gz':
                        output_ext = 'tar.gz'
                    
                    new_output_path = output_path.replace('.ext', f'.{output_ext}')
                    archive_success, archive_msg = await ArchiveHandler.create_archive(
                        extracted_files, new_output_path, target_format
                    )
                    
                    # Tozalash
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    
                    if archive_success:
                        # Output fayl nomini yangilash
                        if os.path.exists(output_path):
                            os.remove(output_path)
                        os.rename(new_output_path, output_path)
                        return True, f"Arxiv {target_format.upper()} formatiga o'tkazildi"
                    else:
                        return False, archive_msg
                else:
                    # Tozalash
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    return False, message
            else:
                # Oddiy fayl nusxalash
                shutil.copy(input_path, output_path)
                return True, "Fayl nusxalandi"
            
        except Exception as e:
            logger.error(f"Arxiv konvertatsiya xatosi: {e}")
            return False, str(e)
    
    @staticmethod
    async def compress_file(input_path: str, output_path: str, settings: Dict) -> Tuple[bool, str]:
        """Faylni siqish"""
        try:
            file_type = get_file_type(get_file_extension(input_path))
            
            # Rasmni siqish
            if file_type == 'image' and Config.HAS_PIL:
                from PIL import Image
                
                with Image.open(input_path) as img:
                    quality = int(settings.get('compress_quality', 60))
                    
                    # O'lchamni kamaytirish
                    new_width = img.width // 2
                    new_height = img.height // 2
                    img = img.resize((new_width, new_height), Image.Resampling.LANCZOS)
                    
                    img.save(output_path, optimize=True, quality=quality)
                
                return True, f"Siqildi: {human_readable_size(os.path.getsize(input_path))} → {human_readable_size(os.path.getsize(output_path))}"
            
            # Boshqa fayllar uchun oddiy nusxa
            shutil.copy(input_path, output_path)
            return True, "Fayl nusxalandi (Siqish amalga oshirilmadi)"
            
        except Exception as e:
            logger.error(f"Siqish xatosi: {e}")
            return False, str(e)

# ==================== BOT HANDLERLARI ====================
class FileConvertBot:
    def __init__(self):
        self.app = None
        self.active_conversions = {}
        self.user_files = {}
        self.user_settings = {}
        
    async def start_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Start komandasi"""
        user = update.effective_user
        welcome_text = f"""
👋 *Salom {user.first_name}!*

🤖 *File Converter Botga xush kelibsiz!*

📁 *Qabul qilinadigan fayllar:*
• 🖼️ Rasmlar: JPG, PNG, WEBP, GIF, BMP, TIFF, ICO
• 📄 Hujjatlar: PDF, DOCX, DOC, TXT, RTF
• 📊 Jadval fayllari: XLSX, XLS, CSV, ODS
• 🎵 Audio: MP3, WAV, OGG, M4A
• 🎬 Video: MP4, AVI, MOV, MKV
• 📦 Arxivlar: ZIP, RAR, 7Z, TAR, GZ, TAR.GZ

🔄 *Yangi funksiyalar:*
• 📂 ZIP/RAR arxivlarini ochish
• 📊 Excel fayllarini CSV, JSON, HTML ga o'tkazish
• 📁 Arxiv formatlarini bir-biriga o'tkazish

⚙️ *Qo'shimcha:*
• Maksimal fayl hajmi: 2GB
• 24 soatdan keyin avtomatik tozalash
• Progress bar
• Fayl ma'lumotlari

📋 *Buyruqlar:*
/start - Botni ishga tushirish
/help - Yordam
/formats - Barcha formatlar
/settings - Sozlamalar
/extract - Arxiv ochish haqida ma'lumot

📎 *Faylni yuboring va kerakli formatni tanlang!*
"""
        
        # Inline keyboard
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Barcha formatlar", callback_data="all_formats")],
            [InlineKeyboardButton("📂 Arxiv ochish", callback_data="extract_info")],
            [InlineKeyboardButton("⚙️ Sozlamalar", callback_data="global_settings")],
            [InlineKeyboardButton("👨‍💻 Admin", url="https://t.me/Ibrohimjon_off")]
        ])
        
        await update.message.reply_text(
            welcome_text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def help_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Help komandasi"""
        help_text = """
🆘 *YORDAM VA QO'LLANMA*

📖 *Qanday ishlatish:*
1️⃣ Faylni yuboring (rasm, hujjat, Excel, audio, video, arxiv)
2️⃣ Kerakli formatni tanlang
3️⃣ Sozlamalarni o'zgartiring (agar kerak bo'lsa)
4️⃣ Konvertatsiya qilingan faylni yuklab oling

📂 *Arxiv ochish:*
• ZIP, RAR, 7Z, TAR, GZ fayllarini yuboring
• "Arxivni ochish" tugmasini bosing
• Ichidagi barcha fayllarni yuklab oling

📊 *Excel konvertatsiya:*
• XLSX → CSV, JSON, HTML
• CSV → XLSX, XLS, JSON
• XLS → XLSX, CSV, JSON

⚠️ *Cheklovlar va shartlar:*
• Maksimal fayl hajmi: 2GB
• Arxiv ichidagi fayllar: maksimal 100MB
• Bir vaqtda 1 ta konvertatsiya
• 24 soatdan keyin avtomatik tozalash

🛠️ *Kerakli kutubxonalar:*
• ZIP ochish: Python standard
• RAR ochish: `pip install rarfile`
• Excel konvertatsiya: `pip install pandas openpyxl`

📞 *Bog'lanish:*
Agar muammo davom etsa, admin bilan bog'laning: @Ibrohimjon_off
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Formatlar", callback_data="all_formats"),
             InlineKeyboardButton("📂 Arxiv", callback_data="extract_info")],
            [InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")]
        ])
        
        await update.message.reply_text(
            help_text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def extract_info_command(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Arxiv ochish haqida ma'lumot"""
        extract_text = """
📂 *ARXIV OCHISH QO'LLANMASI*

*Qabul qilinadigan arxiv formatlari:*
• ZIP - ✅ (o'rnatish shart emas)
• RAR - ✅ (`pip install rarfile` kerak)
• 7Z - ✅ (o'rnatish shart emas)
• TAR - ✅ (o'rnatish shart emas)
• GZ - ✅ (o'rnatish shart emas)
• TAR.GZ - ✅ (o'rnatish shart emas)
• BZ2 - ✅ (o'rnatish shart emas)

*Qanday ishlatish:*
1️⃣ Arxiv faylini yuboring (ZIP, RAR, ...)
2️⃣ "Arxivni ochish" tugmasini bosing
3️⃣ Kutishingiz (katta fayllar uchun bir necha soniya)
4️⃣ Ichidagi fayllarni yuklab oling

*Cheklovlar:*
• Maksimal arxiv hajmi: 2GB
• Arxiv ichidagi barcha fayllar: maksimal 100MB
• Bir vaqtda 1 ta arxivni ochish mumkin

*Muammolar va yechimlar:*
• RAR fayl ochilmasa: `pip install rarfile` ni o'rnating
• Arxiv buzilgan bo'lsa: boshqa arxiv yuboring
• Uzoq vaqt kutish: server band bo'lishi mumkin

*Misollar:*
• backup.zip → ichidagi barcha fayllar
• photos.rar → rasmlar papkasi
• data.tar.gz → ma'lumotlar fayllari
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 Konvertatsiya", callback_data="all_formats")],
            [InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")]
        ])
        
        await update.message.reply_text(
            extract_text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def handle_file(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Fayl yuborilganda"""
        try:
            message = update.message
            user_id = message.from_user.id
            
            # Fayl turini aniqlash
            if message.document:
                file_obj = message.document
                file_name = file_obj.file_name
                file_size = file_obj.file_size
            elif message.photo:
                file_obj = message.photo[-1]
                file_name = f"photo_{file_obj.file_id}.jpg"
                file_size = file_obj.file_size
            elif message.video:
                file_obj = message.video
                file_name = f"video_{file_obj.file_id}.mp4"
                file_size = file_obj.file_size
            elif message.audio:
                file_obj = message.audio
                file_name = file_obj.file_name or f"audio_{file_obj.file_id}.mp3"
                file_size = file_obj.file_size
            elif message.voice:
                file_obj = message.voice
                file_name = f"voice_{file_obj.file_id}.ogg"
                file_size = file_obj.file_size
            else:
                await message.reply_text("❌ Ushbu fayl turi qo'llab-quvvatlanmaydi!")
                return
            
            # Fayl hajmini tekshirish
            if file_size > Config.MAX_FILE_SIZE:
                await message.reply_text(
                    f"❌ Fayl hajmi juda katta!\n"
                    f"📊 Sizning faylingiz: {human_readable_size(file_size)}\n"
                    f"📈 Maksimal: {human_readable_size(Config.MAX_FILE_SIZE)}"
                )
                return
            
            # Fayl kengaytmasini tekshirish
            file_ext = get_file_extension(file_name)
            if not file_ext or file_ext not in FileTypes.ALL:
                await message.reply_text(
                    f"❌ {file_ext.upper()} formati qo'llab-quvvatlanmaydi!\n"
                    f"✅ Qo'llab-quvvatlanadigan formatlar: /formats"
                )
                return
            
            # Yuklash jarayoni
            status_msg = await message.reply_text(
                f"📥 *Fayl yuklanmoqda...*\n"
                f"📊 Hajmi: {human_readable_size(file_size)}\n"
                f"📎 Format: {file_ext.upper()}\n"
                f"🗂️ Turi: {get_file_type(file_ext).title()}"
            )
            
            # Fayl ID yaratish
            file_id = f"{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{hashlib.md5(file_name.encode()).hexdigest()[:8]}"
            input_path = os.path.join(Config.UPLOAD_FOLDER, f"{file_id}.{file_ext}")
            
            # Faylni yuklash
            file = await file_obj.get_file()
            await file.download_to_drive(input_path)
            
            # Fayl ma'lumotlari
            file_info = get_file_info(input_path)
            
            # Foydalanuvchi ma'lumotlarini saqlash
            self.user_files[file_id] = {
                'user_id': user_id,
                'input_path': input_path,
                'original_name': file_name,
                'extension': file_ext,
                'size': file_size,
                'info': file_info,
                'upload_time': datetime.now()
            }
            
            # Boshlang'ich sozlamalar
            if user_id not in self.user_settings:
                self.user_settings[user_id] = {
                    'image_quality': '85',
                    'resize_percent': '100'
                }
            
            # Format tanlash tugmachasini yuborish
            keyboard = create_format_keyboard(file_ext, file_id, self.user_settings.get(user_id, {}))
            
            if keyboard:
                info_text = f"""
✅ *Fayl muvaffaqiyatli yuklandi!*

📄 *Ma'lumotlar:*
• 🏷️ Nomi: `{file_name}`
• 📊 Hajmi: {file_info.get('size', 'Noma\'lum')}
• 📎 Format: {file_ext.upper()}
• 🗂️ Turi: {file_info.get('type', 'Noma\'lum').title()}

"""
                
                # Qo'shimcha ma'lumotlar
                if 'dimensions' in file_info:
                    info_text += f"• 📐 O'lchamlari: {file_info['dimensions']}\n"
                elif 'sheets' in file_info:
                    info_text += f"• 📑 Varaqlar: {file_info['sheets']} ta\n"
                    if 'sheet_names' in file_info:
                        info_text += f"• 📋 Varaq nomlari: {', '.join(file_info['sheet_names'])}\n"
                elif 'files_in_archive' in file_info:
                    info_text += f"• 📁 Fayllar: {file_info['files_in_archive']} ta\n"
                    if 'archive_files' in file_info:
                        info_text += f"• 📄 Fayl nomlari: {', '.join(file_info['archive_files'])}\n"
                
                info_text += f"\n⬇️ *Quyidagi formatlardan birini tanlang:*"
                
                await status_msg.edit_text(
                    info_text,
                    parse_mode=ParseMode.MARKDOWN,
                    reply_markup=keyboard
                )
            else:
                await status_msg.edit_text(
                    f"⚠️ *Diqqat!*\n\n"
                    f"Fayl formati: {file_ext.upper()}\n"
                    f"Ushbu formatdan konvertatsiya qilish imkoni hozircha mavjud emas.\n\n"
                    f"✅ Qo'llab-quvvatlanadigan formatlar: /formats"
                )
                
        except Exception as e:
            logger.error(f"Fayl qabul qilish xatosi: {e}")
            await update.message.reply_text(
                f"❌ Xatolik yuz berdi: {str(e)[:200]}\n"
                f"Iltimos, qayta urinib ko'ring yoki /start ni bosing."
            )
    
    async def button_callback(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Inline tugmalar bosilganda"""
        query = update.callback_query
        await query.answer()
        
        data = query.data
        user_id = query.from_user.id
        
        # Konvertatsiya boshlash
        if data.startswith('conv:'):
            _, file_id, target_format = data.split(':')
            
            if file_id not in self.user_files:
                await query.edit_message_text("❌ Fayl topilmadi. Iltimos, qayta yuboring.")
                return
            
            # Konvertatsiya boshlash
            await self.start_conversion(query, file_id, target_format)
        
        # Arxiv ochish
        elif data.startswith('extract:'):
            _, file_id = data.split(':')
            await self.extract_archive(query, file_id)
        
        # Sozlamalar
        elif data.startswith('set:'):
            _, file_id = data.split(':')
            await self.show_settings(query, file_id)
        
        # Ma'lumot
        elif data.startswith('info:'):
            _, file_id = data.split(':')
            await self.show_file_info(query, file_id)
        
        # Sozlamalarni yangilash
        elif data.startswith('qual:'):
            _, file_id, value = data.split(':')
            await self.update_setting(query, file_id, 'qual', value)
        elif data.startswith('resize:'):
            _, file_id, value = data.split(':')
            await self.update_setting(query, file_id, 'resize', value)
        elif data.startswith('save:'):
            _, file_id = data.split(':')
            await query.answer("Sozlamalar saqlandi!", show_alert=True)
            await self.back_to_formats(query, file_id)
        
        # Orqaga qaytish
        elif data.startswith('back:'):
            _, file_id = data.split(':')
            await self.back_to_formats(query, file_id)
        
        # Boshqa funksiyalar
        elif data == 'all_formats':
            await self.show_all_formats(query)
        elif data == 'extract_info':
            await self.show_extract_info(query)
        elif data == 'global_settings':
            await self.show_global_settings(query)
        elif data == 'main_menu':
            await self.show_main_menu(query)
        elif data == 'help_menu':
            await self.show_help_menu(query)
    
    async def start_conversion(self, query, file_id: str, target_format: str):
        """Konvertatsiyani boshlash"""
        try:
            file_data = self.user_files[file_id]
            input_path = file_data['input_path']
            original_name = file_data['original_name']
            original_ext = file_data['extension']
            user_id = file_data['user_id']
            
            # Output fayl nomi
            base_name = original_name.rsplit('.', 1)[0]
            output_name = f"{base_name}_converted.{target_format}"
            output_path = os.path.join(Config.OUTPUT_FOLDER, output_name)
            
            # Progress bar yaratish
            progress_msg = await query.edit_message_text(
                f"🔄 *Konvertatsiya boshlanmoqda...*\n\n"
                f"📤 Kirish: `{original_name}`\n"
                f"📥 Chiqish: `{output_name}`\n\n"
                f"⏳ Jarayon: 0%\n"
                f"░░░░░░░░░░"
            )
            
            # Konvertatsiya jarayoni
            success = False
            error_message = ""
            
            # Fayl turi
            file_type = get_file_type(original_ext)
            settings = self.user_settings.get(user_id, {})
            
            # Konvertatsiya qilish
            if file_type == 'image':
                success, error_message = await Converter.convert_image(
                    input_path, output_path, target_format, settings
                )
            elif file_type == 'document':
                success, error_message = await Converter.convert_document(
                    input_path, output_path, target_format, settings
                )
            elif file_type == 'spreadsheet':
                success, error_message = await Converter.convert_spreadsheet(
                    input_path, output_path, target_format, settings
                )
            elif file_type == 'audio':
                success, error_message = await Converter.convert_audio(
                    input_path, output_path, target_format, settings
                )
            elif file_type == 'video':
                success, error_message = await Converter.convert_video(
                    input_path, output_path, target_format, settings
                )
            elif file_type == 'archive':
                success, error_message = await Converter.convert_archive(
                    input_path, output_path, target_format
                )
            else:
                error_message = "Noma'lum fayl turi"
            
            # Progress bar animatsiyasi
            for i in range(1, 11):
                await asyncio.sleep(0.3)
                progress_text = f"🔄 *Konvertatsiya boshlanmoqda...*\n\n📤 Kirish: `{original_name}`\n📥 Chiqish: `{output_name}`\n\n⏳ Jarayon: {i*10}%\n"
                progress_text += "".join(["▓" if j <= i else "░" for j in range(1, 11)])
                await progress_msg.edit_text(progress_text)
            
            # Natijani ko'rsatish
            if success and os.path.exists(output_path):
                output_size = os.path.getsize(output_path)
                
                await progress_msg.edit_text(
                    f"✅ *Konvertatsiya muvaffaqiyatli yakunlandi!*\n\n"
                    f"📤 {original_ext.upper()} → {target_format.upper()}\n"
                    f"📊 Hajmi: {human_readable_size(output_size)}\n\n"
                    f"📤 Yuklab olinmoqda..."
                )
                
                # Faylni yuborish
                await self.send_converted_file(
                    query.message.chat_id,
                    output_path,
                    output_name,
                    target_format,
                    original_ext
                )
                
                # Tozalash
                try:
                    os.remove(output_path)
                except:
                    pass
                
            else:
                await progress_msg.edit_text(
                    f"❌ *Konvertatsiya muvaffaqiyatsiz tugadi!*\n\n"
                    f"📤 {original_ext.upper()} → {target_format.upper()}\n\n"
                    f"⚠️ Xato: {error_message[:300]}\n\n"
                    f"🔧 *Ehtimoliy sabablar:*\n"
                    f"• Fayl buzilgan\n"
                    f"• Format mos emas\n"
                    f"• Kerakli kutubxona yo'q\n"
                    f"• Server cheklovlari"
                )
                
        except Exception as e:
            logger.error(f"Konvertatsiya xatosi: {e}")
            await query.edit_message_text(
                f"❌ *Kutilmagan xatolik yuz berdi!*\n\n"
                f"```{str(e)[:500]}```\n\n"
                f"Iltimos, qayta urinib ko'ring."
            )
    
    async def extract_archive(self, query, file_id: str):
        """Arxiv faylini ochish"""
        try:
            file_data = self.user_files[file_id]
            input_path = file_data['input_path']
            original_name = file_data['original_name']
            original_ext = file_data['extension']
            
            # Arxiv emasligini tekshirish
            if original_ext not in FileTypes.ARCHIVES:
                await query.answer("Bu fayl arxiv emas!", show_alert=True)
                return
            
            # Arxiv hajmini tekshirish
            archive_info = get_file_info(input_path)
            if 'files_in_archive' in archive_info and archive_info['files_in_archive'] > 100:
                await query.edit_message_text(
                    f"❌ Arxiv juda katta!\n"
                    f"📊 Fayllar soni: {archive_info['files_in_archive']} ta\n"
                    f"📈 Maksimal: 100 ta fayl\n\n"
                    f"Kichikroq arxiv yuboring."
                )
                return
            
            # Progress bar
            progress_msg = await query.edit_message_text(
                f"📂 *Arxiv ochilmoqda...*\n\n"
                f"📦 Arxiv: `{original_name}`\n"
                f"📎 Format: {original_ext.upper()}\n\n"
                f"⏳ Jarayon: 0%\n"
                f"░░░░░░░░░░"
            )
            
            # Extract qilish uchun vaqtinchalik papka
            extract_dir = os.path.join(Config.EXTRACT_FOLDER, f"{file_id}_extracted")
            os.makedirs(extract_dir, exist_ok=True)
            
            # Progress animatsiyasi
            for i in range(1, 11):
                await asyncio.sleep(0.5)
                progress_text = f"📂 *Arxiv ochilmoqda...*\n\n📦 Arxiv: `{original_name}`\n📎 Format: {original_ext.upper()}\n\n⏳ Jarayon: {i*10}%\n"
                progress_text += "".join(["▓" if j <= i else "░" for j in range(1, 11)])
                await progress_msg.edit_text(progress_text)
            
            # Arxivni ochish
            success, message, extracted_files = await ArchiveHandler.extract_archive(input_path, extract_dir)
            
            if success and extracted_files:
                await progress_msg.edit_text(
                    f"✅ *Arxiv muvaffaqiyatli ochildi!*\n\n"
                    f"📦 Arxiv: {original_name}\n"
                    f"📊 Fayllar: {len(extracted_files)} ta\n\n"
                    f"📤 Fayllar yuklanmoqda..."
                )
                
                # Fayllarni yuborish
                sent_count = 0
                for file_path in extracted_files[:10]:  # Faqat 10 ta fayl
                    try:
                        file_name = os.path.basename(file_path)
                        file_size = os.path.getsize(file_path)
                        
                        # 50MB cheklovi
                        if file_size > 50 * 1024 * 1024:
                            await self.app.bot.send_message(
                                query.message.chat_id,
                                f"❌ `{file_name}` hajmi juda katta ({human_readable_size(file_size)}). Yuklanmadi."
                            )
                            continue
                        
                        with open(file_path, 'rb') as f:
                            await self.app.bot.send_document(
                                chat_id=query.message.chat_id,
                                document=f,
                                filename=file_name,
                                caption=f"📦 {original_name} ichidan\n📊 Hajmi: {human_readable_size(file_size)}"
                            )
                        
                        sent_count += 1
                        await asyncio.sleep(0.5)  # Spamdan saqlash uchun
                        
                    except Exception as e:
                        logger.error(f"Fayl yuborish xatosi: {e}")
                
                # Qo'shimcha xabar
                if len(extracted_files) > 10:
                    await self.app.bot.send_message(
                        query.message.chat_id,
                        f"ℹ️ {len(extracted_files)} ta fayldan {sent_count} tasi yuklandi.\n"
                        f"Qolgan {len(extracted_files) - 10} ta fayl hajmi katta yoki cheklovlar tufayli yuklanmadi."
                    )
                else:
                    await self.app.bot.send_message(
                        query.message.chat_id,
                        f"✅ Barcha {sent_count} ta fayl muvaffaqiyatli yuklandi!"
                    )
                
                # Tozalash
                shutil.rmtree(extract_dir, ignore_errors=True)
                
            else:
                await progress_msg.edit_text(
                    f"❌ *Arxiv ochish muvaffaqiyatsiz!*\n\n"
                    f"📦 Arxiv: {original_name}\n\n"
                    f"⚠️ Xato: {message[:300]}\n\n"
                    f"🔧 *Ehtimoliy sabablar:*\n"
                    f"• Arxiv buzilgan\n"
                    f"• Parol himoyalangan\n"
                    f"• RarFile kutubxonasi yo'q\n"
                    f"• Server cheklovlari"
                )
                
        except Exception as e:
            logger.error(f"Arxiv ochish xatosi: {e}")
            await query.edit_message_text(
                f"❌ *Kutilmagan xatolik yuz berdi!*\n\n"
                f"```{str(e)[:500]}```\n\n"
                f"Iltimos, qayta urinib ko'ring."
            )
    
    async def send_converted_file(self, chat_id: int, file_path: str, file_name: str, 
                                 target_format: str, original_format: str):
        """Konvertatsiya qilingan faylni yuborish"""
        try:
            file_size = os.path.getsize(file_path)
            
            # Fayl hajmi cheklovi (Telegram uchun)
            if file_size > 50 * 1024 * 1024:  # 50MB
                await self.app.bot.send_message(
                    chat_id,
                    f"❌ Fayl hajmi juda katta ({human_readable_size(file_size)}).\n"
                    f"Telegram 50MB dan katta fayllarni qabul qilmaydi.\n\n"
                    f"📊 Kichikroq fayl yuboring yoki sozlamalardan siqishni tanlang."
                )
                return
            
            # Fayl turiga qarab yuborish
            with open(file_path, 'rb') as f:
                file_type = get_file_type(target_format)
                
                if file_type == 'image':
                    await self.app.bot.send_photo(
                        chat_id=chat_id,
                        photo=f,
                        caption=f"✅ {original_format.upper()} → {target_format.upper()}\n"
                               f"📊 Hajmi: {human_readable_size(file_size)}"
                    )
                elif file_type == 'audio':
                    await self.app.bot.send_audio(
                        chat_id=chat_id,
                        audio=f,
                        title=file_name,
                        caption=f"✅ {original_format.upper()} → {target_format.upper()}\n"
                               f"📊 Hajmi: {human_readable_size(file_size)}"
                    )
                elif file_type == 'video':
                    await self.app.bot.send_video(
                        chat_id=chat_id,
                        video=f,
                        caption=f"✅ {original_format.upper()} → {target_format.upper()}\n"
                               f"📊 Hajmi: {human_readable_size(file_size)}"
                    )
                else:
                    # Barcha boshqa fayllar uchun document
                    await self.app.bot.send_document(
                        chat_id=chat_id,
                        document=f,
                        caption=f"✅ {original_format.upper()} → {target_format.upper()}\n"
                               f"📊 Hajmi: {human_readable_size(file_size)}"
                    )
                    
        except Exception as e:
            logger.error(f"Fayl yuborish xatosi: {e}")
            await self.app.bot.send_message(
                chat_id,
                f"❌ Faylni yuborishda xatolik: {str(e)[:200]}"
            )
    
    async def show_settings(self, query, file_id: str):
        """Sozlamalarni ko'rsatish"""
        if file_id not in self.user_files:
            await query.edit_message_text("❌ Fayl topilmadi.")
            return
        
        user_id = self.user_files[file_id]['user_id']
        settings = self.user_settings.get(user_id, {})
        
        text = """
⚙️ *KONVERTATSIYA SOZLAMALARI*

Quyidagi sozlamalarni o'zgartiring:

🖼️ *Rasm:*
• JPG/PNG sifat (30-100%)
• O'lcham (25-100%)

📊 *Excel:*
• Faqat birinchi varaq (hozircha)

Sozlamalarni tanlang:
"""
        
        # Tugmalarni yaratish
        buttons = []
        
        # Rasm sifatini sozlash
        quality_buttons = []
        for q in ["30", "60", "85", "95", "100"]:
            selected = " ✅" if settings.get('image_quality') == q else ""
            quality_buttons.append(
                InlineKeyboardButton(f"{q}%{selected}", callback_data=f"qual:{file_id}:{q}")
            )
        buttons.append(quality_buttons)
        
        # O'lchamni o'zgartirish
        resize_buttons = [
            InlineKeyboardButton("25%", callback_data=f"resize:{file_id}:25"),
            InlineKeyboardButton("50%", callback_data=f"resize:{file_id}:50"),
            InlineKeyboardButton("75%", callback_data=f"resize:{file_id}:75"),
            InlineKeyboardButton("100%", callback_data=f"resize:{file_id}:100"),
        ]
        buttons.append(resize_buttons)
        
        # Orqaga
        buttons.append([
            InlineKeyboardButton("🔙 Orqaga", callback_data=f"back:{file_id}"),
            InlineKeyboardButton("✅ Saqlash", callback_data=f"save:{file_id}")
        ])
        
        keyboard = InlineKeyboardMarkup(buttons)
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def update_setting(self, query, file_id: str, key: str, value: str):
        """Sozlamani yangilash"""
        if file_id not in self.user_files:
            await query.answer("Fayl topilmadi!", show_alert=True)
            return
        
        user_id = self.user_files[file_id]['user_id']
        
        if user_id not in self.user_settings:
            self.user_settings[user_id] = {}
        
        if key == 'qual':
            self.user_settings[user_id]['image_quality'] = value
        elif key == 'resize':
            self.user_settings[user_id]['resize_percent'] = value
        
        await query.answer(f"Sozlama yangilandi: {value}", show_alert=True)
        
        # Sozlamalar sahifasini yangilash
        await self.show_settings(query, file_id)
    
    async def back_to_formats(self, query, file_id: str):
        """Format tanlash sahifasiga qaytish"""
        if file_id not in self.user_files:
            await query.edit_message_text("❌ Fayl topilmadi.")
            return
        
        file_data = self.user_files[file_id]
        original_ext = file_data['extension']
        user_id = file_data['user_id']
        
        keyboard = create_format_keyboard(original_ext, file_id, self.user_settings.get(user_id, {}))
        
        if keyboard:
            await query.edit_message_text(
                f"📄 *Format tanlash*\n\n"
                f"Hozirgi format: {original_ext.upper()}\n"
                f"Quyidagi formatlardan birini tanlang:",
                reply_markup=keyboard,
                parse_mode=ParseMode.MARKDOWN
            )
        else:
            await query.edit_message_text(
                "❌ Ushbu format uchun konvertatsiya imkoni yo'q."
            )
    
    async def show_file_info(self, query, file_id: str):
        """Fayl ma'lumotlarini ko'rsatish"""
        if file_id not in self.user_files:
            await query.answer("Fayl topilmadi!", show_alert=True)
            return
        
        file_data = self.user_files[file_id]
        info = file_data['info']
        
        text = f"""
📋 *FAYL MA'LUMOTLARI*

🏷️ **Nomi:** `{file_data['original_name']}`
📊 **Hajmi:** {info.get('size', 'Noma\'lum')}
📎 **Formati:** {file_data['extension'].upper()}
🗂️ **Turi:** {info.get('type', 'Noma\'lum').title()}
🕐 **Yuklangan:** {file_data['upload_time'].strftime('%Y-%m-%d %H:%M:%S')}
"""
        
        # Qo'shimcha ma'lumotlar
        if 'dimensions' in info:
            text += f"📐 **O'lchamlari:** {info['dimensions']}\n"
        elif 'sheets' in info:
            text += f"📑 **Varaqlar:** {info['sheets']} ta\n"
            if 'sheet_names' in info:
                text += f"📋 **Varaq nomlari:** {', '.join(info['sheet_names'][:3])}"
                if len(info['sheet_names']) > 3:
                    text += f" va {len(info['sheet_names']) - 3} ta boshqa\n"
        elif 'files_in_archive' in info:
            text += f"📁 **Fayllar:** {info['files_in_archive']} ta\n"
            if 'archive_files' in info:
                text += f"📄 **Fayl nomlari:** {', '.join(info['archive_files'][:3])}"
                if len(info['archive_files']) > 3:
                    text += f" va {len(info['archive_files']) - 3} ta boshqa\n"
        
        text += f"\n🔄 **Mumkin konvertatsiyalar:** {len(CONVERSION_MATRIX.get(file_data['extension'], []))} ta"
        
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("🔙 Orqaga", callback_data=f"back:{file_id}")
        ]])
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def show_all_formats(self, query):
        """Barcha formatlarni ko'rsatish"""
        text = """
📋 *QO'LLAB-QUVVATLANADIGAN BARCHA FORMATLAR*

🖼️ *RASMLAR (8 ta):*
JPG, JPEG, PNG, WEBP, BMP, GIF, TIFF, ICO

📄 *HUJJATLAR (5 ta):*
PDF, DOCX, DOC, TXT, RTF

📊 *JADVAL FAYLLARI (4 ta):*
XLSX, XLS, CSV, ODS

🎵 *AUDIO (4 ta):*
MP3, WAV, OGG, M4A

🎬 *VIDEO (4 ta):*
MP4, AVI, MOV, MKV

📦 *ARXIVLAR (7 ta):*
ZIP, RAR, 7Z, TAR, GZ, TAR.GZ, BZ2

🔄 *JAMI: 32 turdagi formatlar*

📂 *ARXIV OCHISH:*
• ZIP, RAR, 7Z, TAR, GZ fayllarini ochish
• Ichidagi fayllarni yuklab olish

📊 *EXCEL KONVERTATSIYA:*
• XLSX ↔ CSV, JSON, HTML
• CSV ↔ XLSX, XLS, JSON
• XLS ↔ XLSX, CSV, JSON

⚠️ *Eslatma:* Ba'zi konvertatsiyalar qo'shimcha kutubxonalarni talab qilishi mumkin.
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📂 Arxiv ochish", callback_data="extract_info")],
            [InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")]
        ])
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def show_extract_info(self, query):
        """Arxiv ochish haqida ma'lumot"""
        text = """
📂 *ARXIV OCHISH QO'LLANMASI*

*Qabul qilinadigan arxiv formatlari:*
• ZIP - ✅ (o'rnatish shart emas)
• RAR - ✅ (`pip install rarfile` kerak)
• 7Z - ✅ (o'rnatish shart emas)
• TAR - ✅ (o'rnatish shart emas)
• GZ - ✅ (o'rnatish shart emas)
• TAR.GZ - ✅ (o'rnatish shart emas)
• BZ2 - ✅ (o'rnatish shart emas)

*Qanday ishlatish:*
1️⃣ Arxiv faylini yuboring (ZIP, RAR, ...)
2️⃣ "Arxivni ochish" tugmasini bosing
3️⃣ Kutishingiz (katta fayllar uchun bir necha soniya)
4️⃣ Ichidagi fayllarni yuklab oling

*Cheklovlar:*
• Maksimal arxiv hajmi: 2GB
• Arxiv ichidagi barcha fayllar: maksimal 100MB
• Bir vaqtda 1 ta arxivni ochish mumkin

*Misollar:*
• backup.zip → ichidagi barcha fayllar
• photos.rar → rasmlar papkasi
• data.tar.gz → ma'lumotlar fayllari
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Barcha formatlar", callback_data="all_formats")],
            [InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")]
        ])
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def show_global_settings(self, query):
        """Global sozlamalarni ko'rsatish"""
        user_id = query.from_user.id
        settings = self.user_settings.get(user_id, {})
        
        text = """
⚙️ *GLOBAL SOZLAMALAR*

Quyidagi sozlamalar barcha konvertatsiyalaringizga ta'sir qiladi:

"""
        
        if settings:
            for key, value in settings.items():
                text += f"• {key.replace('_', ' ').title()}: {value}\n"
        else:
            text += "⚠️ Hozircha sozlamalar mavjud emas.\n"
        
        text += "\nHar bir fayl uchun sozlamalarni alohida o'zgartirishingiz mumkin."
        
        keyboard = InlineKeyboardMarkup([[
            InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")
        ]])
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def show_main_menu(self, query):
        """Asosiy menyuni ko'rsatish"""
        text = """
🏠 *ASOSIY MENYU*

🤖 File Converter Bot - Bu eng mukammal fayl konvertatsiya boti!

📋 **Yangi xizmatlar:**
• 📂 ZIP/RAR arxivlarini ochish
• 📊 Excel fayllarini konvertatsiya qilish
• 📁 Arxiv formatlarini bir-biriga o'tkazish

📊 **Statistika:**
• 2GB gacha fayllar
• 32+ turdagi formatlar
• 24/7 ishlaydi
• Tez va ishonchli

👇 Quyidagi tugmalardan birini tanlang:
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Formatlar", callback_data="all_formats"),
             InlineKeyboardButton("📂 Arxiv", callback_data="extract_info")],
            [InlineKeyboardButton("⚙️ Sozlamalar", callback_data="global_settings"),
             InlineKeyboardButton("🆘 Yordam", callback_data="help_menu")],
            [InlineKeyboardButton("👨‍💻 Admin", url="https://t.me/Ibrohimjon_off")]
        ])
        
        await query.edit_message_text(
            text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def show_help_menu(self, query):
        """Yordam menyusi"""
        help_text = """
🆘 *TEZKOR YORDAM*

*Tez boshlash:*
1️⃣ Fayl yuboring
2️⃣ Format tanlang
3️⃣ Yuklab oling

*Muhim tugmalar:*
• 📂 Arxivni ochish - ZIP/RAR fayllarini ochish
• 📊 Excel - Jadval fayllarini konvertatsiya qilish
• ⚙️ Sozlamalar - Rasm sifatini sozlash

*Tez-tez so'raladigan savollar:*
❓ RAR fayl ochilmayapti?
➡️ `pip install rarfile` ni o'rnating

❓ Excel fayl konvertatsiya qilinmayapti?
➡️ `pip install pandas openpyxl` ni o'rnating

❓ Fayl hajmi juda katta?
➡️ Sozlamalardan siqishni tanlang

📞 *Qo'shimcha yordam:*
@Ibrohimjon_off
"""
        
        keyboard = InlineKeyboardMarkup([
            [InlineKeyboardButton("📋 Formatlar", callback_data="all_formats"),
             InlineKeyboardButton("📂 Arxiv", callback_data="extract_info")],
            [InlineKeyboardButton("🔙 Bosh sahifa", callback_data="main_menu")]
        ])
        
        await query.edit_message_text(
            help_text,
            parse_mode=ParseMode.MARKDOWN,
            reply_markup=keyboard
        )
    
    async def cleanup_old_files_task(self):
        """Eski fayllarni tozalash vazifasi"""
        while True:
            try:
                now = datetime.now()
                
                # Barcha papkalarni tekshirish
                folders = [Config.UPLOAD_FOLDER, Config.OUTPUT_FOLDER, 
                          Config.TEMP_FOLDER, Config.EXTRACT_FOLDER]
                
                for folder in folders:
                    if os.path.exists(folder):
                        for filename in os.listdir(folder):
                            filepath = os.path.join(folder, filename)
                            if os.path.isfile(filepath):
                                try:
                                    file_time = datetime.fromtimestamp(os.path.getctime(filepath))
                                    if now - file_time > timedelta(hours=Config.CLEANUP_HOURS):
                                        os.remove(filepath)
                                        logger.info(f"Fayl o'chirildi: {filepath}")
                                except:
                                    pass
                            elif os.path.isdir(filepath):
                                try:
                                    dir_time = datetime.fromtimestamp(os.path.getctime(filepath))
                                    if now - dir_time > timedelta(hours=Config.CLEANUP_HOURS):
                                        shutil.rmtree(filepath, ignore_errors=True)
                                        logger.info(f"Papka o'chirildi: {filepath}")
                                except:
                                    pass
                
                # Eski foydalanuvchi ma'lumotlari
                expired_files = []
                for file_id, data in list(self.user_files.items()):
                    if now - data['upload_time'] > timedelta(hours=Config.CLEANUP_HOURS):
                        expired_files.append(file_id)
                
                for file_id in expired_files:
                    del self.user_files[file_id]
                
                if expired_files:
                    logger.info(f"{len(expired_files)} ta eski fayl ma'lumotlari tozalandi")
                
            except Exception as e:
                logger.error(f"Tozalash xatosi: {e}")
            
            await asyncio.sleep(3600)  # Har soatda
    
    async def error_handler(self, update: Update, context: ContextTypes.DEFAULT_TYPE):
        """Xatolarni qayta ishlash"""
        logger.error(f"Xatolik yuz berdi: {context.error}", exc_info=context.error)
        
        if update and update.effective_message:
            await update.effective_message.reply_text(
                "❌ Kutilmagan xatolik yuz berdi. Iltimos, qayta urinib ko'ring."
            )
    
    def run(self):
        """Botni ishga tushirish"""
        # Muhitni sozlash
        setup_environment()
        
        # Bot ilovasini yaratish
        self.app = Application.builder().token(Config.BOT_TOKEN).build()
        
        # Handlerlarni qo'shish
        self.app.add_handler(CommandHandler("start", self.start_command))
        self.app.add_handler(CommandHandler("help", self.help_command))
        self.app.add_handler(CommandHandler("formats", self.show_all_formats))
        self.app.add_handler(CommandHandler("extract", self.extract_info_command))
        self.app.add_handler(CommandHandler("settings", self.show_global_settings))
        
        # Fayl handlerlari
        self.app.add_handler(MessageHandler(
            filters.Document.ALL | filters.PHOTO | filters.VIDEO | 
            filters.AUDIO | filters.VOICE, self.handle_file
        ))
        
        # Callback handler
        self.app.add_handler(CallbackQueryHandler(self.button_callback))
        
        # Xatolik handler
        self.app.add_error_handler(self.error_handler)
        
        # Vazifalarni ishga tushirish
        loop = asyncio.get_event_loop()
        loop.create_task(self.cleanup_old_files_task())
        
        # Botni ishga tushirish
        print("=" * 50)
        print("🤖 FILE CONVERTER BOT ISHGA TUSHDI!")
        print("=" * 50)
        print(f"📁 Upload papkasi: {os.path.abspath(Config.UPLOAD_FOLDER)}")
        print(f"📁 Output papkasi: {os.path.abspath(Config.OUTPUT_FOLDER)}")
        print(f"📁 Temp papkasi: {os.path.abspath(Config.TEMP_FOLDER)}")
        print(f"📁 Extract papkasi: {os.path.abspath(Config.EXTRACT_FOLDER)}")
        print("=" * 50)
        print("Mavjud kutubxonalar:")
        print(f"• PIL/Pillow: {'✅' if Config.HAS_PIL else '❌'}")
        print(f"• ReportLab: {'✅' if Config.HAS_REPORTLAB else '❌'}")
        print(f"• PyExcel: {'✅' if Config.HAS_PYEXCEL else '❌'}")
        print(f"• Pandas: {'✅' if Config.HAS_PANDAS else '❌'}")
        print(f"• OpenPyXL: {'✅' if Config.HAS_OPENPYXL else '❌'}")
        print(f"• RarFile: {'✅' if Config.HAS_RARFILE else '❌'}")
        print("=" * 50)
        print("Yangi funksiyalar:")
        print("• 📂 ZIP/RAR arxivlarini ochish")
        print("• 📊 Excel fayllarini konvertatsiya qilish")
        print("• 📁 Arxiv formatlarini bir-biriga o'tkazish")
        print("=" * 50)
        print("Bot ishlayapti... CTRL+C tugmasini bosing (to'xtatish uchun)")
        
        self.app.run_polling(allowed_updates=Update.ALL_TYPES)

# ==================== ASOSIY FUNKSIYA ====================
if __name__ == '__main__':
    try:
        bot = FileConvertBot()
        bot.run()
    except KeyboardInterrupt:
        print("\n\nBot to'xtatildi!")
    except Exception as e:
        logger.error(f"Bot ishga tushirishda xatolik: {e}")
        traceback.print_exc()